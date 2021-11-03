import os
import openpyxl
from PIL import Image


def pic_size(pic_name, target):  # 사진 크기 확인
    pic = Image.open(pic_name)
    size = [pic.size[0], pic.size[1]]
    pic.close()
    if target != size[0] or target != size[1]:
        print("{}의 크기가 {} x {} 가 아닙니다".format(file_name, target, target))


def no_pic_add(lst, brand, extension, no_lst):
    no_brand_temp = True  # 사진 없는 브랜드 추가
    for ext in extension:
        if brand + ext in lst:
            no_brand_temp = False
            break
    if no_brand_temp:
        no_lst.append(brand)


def no_pic_list(lst):  # 사진 없는 브랜드 출력
    if len(lst) != 0:
        print("\n다음은 사진이 없는 브랜드입니다. 등록된 브랜드인지 확인해주세요")
        brand_temp = []
        for brand in lst:
            brand_temp.append(brand)
            if len(brand_temp) == 5:
                print(
                    "{} {} {} {} {}".format(brand_temp[0], brand_temp[1], brand_temp[2], brand_temp[3], brand_temp[4]))
                brand_temp = []
        if len(brand_temp) != 0:
            rmd = ''
            for brand in brand_temp:
                if len(rmd) == 0:
                    rmd = brand
                else:
                    rmd = rmd + " " + brand
            print(rmd)


def image_check(lst, brand, img, directory):
    if img[-4:] != img[-4:].lower():  # 확장자명 대문자
        print("{} 파일은 확장자명이 대문자입니다".format(directory))

    if img[:-4] == brand:  # 브랜드사진
        pic_size(directory, 150)
    else:  # 제품사진
        pic_size(directory, 480)
        if img[:-4] + '.xlsx' not in lst:
            print("{} 의 엑셀파일이 없습니다".format(directory))


def product_check(lst, brand, xls, directory, extension, worksheet):
    no_pic_temp = True  # 제품사진부재확인
    for typ in extension:
        if xls[:-5] + typ in lst:
            no_pic_temp = False
            break
    if no_pic_temp:
        print("{} 의 제품 사진이 없습니다".format(directory))

    wb = openpyxl.load_workbook(filename=directory)
    ws = wb[worksheet]

    if ws['C2'].value != xls[:-5]:  # 제품명불일치
        print("{} 의 파일명과 제품명이 다릅니다".format(directory))

    if ws['C3'].value != brand:  # 브랜드명불이치
        print("{} 의 폴더명과 브랜드명이 다릅니다".format(directory))

    if ws['A26'].value == "" or ws['A26'].value == 0 :  # 성분이 없음
        print("{} 의 성분이 없습니다".format(directory))

    for i in range(26, 227):  # -1 찾기, 현재 200개 탐색
        if ws['A' + str(i)].value == -1:
            print("{} 의 성분 {}에 성분번호가 할당되지 않았습니다".format(directory, ws['B' + str(i)].value))
        if (ws['A' + str(i)].value is None or ws['A' + str(i)].value == 0) and (ws['B' + str(i)].value != 0 and ws['B' + str(i)].value is not None):
            print("{} 의 성분 {}에 성분명이 할당되지 않았습니다".format(directory, ws['A' + str(i)].value))
    wb.close()


readme = '''수집 오류 확인기 v1.0.2: 현재 구현된 기능:
제품사진크기부정확 브랜드사진크기부정확 확장자명대문자 확장자명앞스페이스
사진에맞는엑셀부재 엑셀에맞는사진부재 파일명과제품명불일치
폴더명과브랜드명불일치 성분번호없음 브랜드사진없는브랜드
성분명없음 성분없음 신규수집양식오류해결(1.0.2)'''
print(readme)

sht_idx = input("1. liv 2. cos: ")
sht = ['liv', 'cos'][int(sht_idx) - 1]

brand_count = 0
image_count = 0
product_count = 0
no_pic_brand = []
image_extension = ['.jpg', '.png', '.JPG', '.PNG']

folder_path = os.getcwd()
folder_path = folder_path + '\\'
folder_list = os.listdir(folder_path)

for folder in folder_list:  # 브랜드 탐색
    if (folder[-3:] == '.py') or (folder[-4:] == '.bat'):
        continue
    else:
        brand_count = brand_count + 1

        file_path = folder_path + folder
        file_list = os.listdir(file_path)

        no_pic_add(file_list, folder, image_extension, no_pic_brand)

        for file in file_list:  # 파일 탐색
            file_name = file_path + '\\' + file

            if file.find(' .') != -1:  # 파일명 스페이스바
                print("{} 파일은 확장자명 앞에 스페이스가 있습니다".format(file_name))
                
            if file[-5:] != '.xlsx':  # 이미지의 경우
                image_count = image_count + 1
                image_check(file_list, folder, file, file_name)

            else:  # 제품엑셀
                product_count = product_count + 1
                product_check(file_list, folder, file, file_name, image_extension, sht)

no_pic_list(no_pic_brand)

print("\n총 {}개 브랜드 {}개 이미지 {}개 제품 확인 완료하였습니다.".format(brand_count, image_count, product_count))
